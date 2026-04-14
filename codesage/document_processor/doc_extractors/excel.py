"""基于 RAGFlow 轻量流程改造的表格文档提取辅助函数。"""

from __future__ import annotations

from pathlib import Path

from .base import join_sections, table_rows_to_text


def _dataframe_rows(frame) -> list[list[object]]:
    rows = [list(frame.columns)]
    for row in frame.itertuples(index=False, name=None):
        rows.append(list(row))
    return rows


def _extract_csv(filepath: str) -> str:
    import pandas as pd

    frame = pd.read_csv(filepath, on_bad_lines="skip")
    return table_rows_to_text(_dataframe_rows(frame), caption="工作表：数据")


def _extract_with_openpyxl(filepath: str) -> str:
    from openpyxl import load_workbook

    workbook = load_workbook(filepath, data_only=True)
    sections: list[str] = []

    for sheet in workbook.worksheets:
        rows = [
            [cell for cell in row]
            for row in sheet.iter_rows(values_only=True)
            if any(cell not in (None, "") for cell in row)
        ]
        if not rows:
            continue
        sections.append(table_rows_to_text(rows, caption=f"工作表：{sheet.title}"))

    return join_sections(sections)


def _extract_with_pandas(filepath: str) -> str:
    import pandas as pd

    frames = pd.read_excel(filepath, sheet_name=None)
    sections = [
        table_rows_to_text(_dataframe_rows(frame), caption=f"工作表：{sheet_name}")
        for sheet_name, frame in frames.items()
    ]
    return join_sections(sections)


def extract_text(filepath: str) -> str:
    """从 CSV 和 Excel 文件中提取文本。"""
    ext = Path(filepath).suffix.lower()
    if ext == ".csv":
        return join_sections([_extract_csv(filepath)])

    if ext != ".xls":
        try:
            text = _extract_with_openpyxl(filepath)
            if text:
                return text
        except Exception:
            pass

    return _extract_with_pandas(filepath)
